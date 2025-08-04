import matplotlib
matplotlib.use('Agg') # This tells matplotlib to use a non-interactive backend for Streamlit deployment

import streamlit as st
import pdfplumber # <--- NEW import
import pandas as pd
import io
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# --- Helper Functions and Data ---
# (Keep all your existing helper functions and UNIT_CONVERSIONS as they are)

# Define a dictionary for unit conversions
UNIT_CONVERSIONS = {
    "p-no/p hour": 1.0, "p-no": 1.0, "hour": 1.0, "hours": 1.0, "meters": 1.0,
    "kilometers": 1000.0, "each": 1.0, "lot": 1.0, "kgs": 1.0, "tons": 1000.0,
    "liters": 1.0, "gallons": 3.785, "days": 8.0, "nos": 1.0, "pc": 1.0,
    "set": 1.0, "%": 100.0, "%0": 1000.0, "%o": 1000.0, "sqm": 1.0, "cum": 1.0,
    "lm": 1.0,
}

def convert_unit_to_number(unit_string):
    """
    Converts a unit string to a numerical value based on predefined mappings.
    Prioritizes specific percentage logic if '%', '%0', '%o', '% 0', or '% o' are present,
    otherwise uses the UNIT_CONVERSIONS dictionary.
    """
    if not isinstance(unit_string, str):
        return 0.0  # Return 0 for non-string types (e.g., NaN, None)

    unit_string_cleaned = unit_string.strip().lower()

    # Specific logic for percentage units for 1000.0
    if "%0" in unit_string_cleaned or "%o" in unit_string_cleaned or "% 0" in unit_string_cleaned or "% o" in unit_string_cleaned:
        return 1000.0
    elif "%" in unit_string_cleaned:
        return 100.0

    return UNIT_CONVERSIONS.get(unit_string_cleaned, 1.0)

def calculate_total_rate(input_rate, quantity, units_no):
    """Calculates Total Rate: (Input Rate * Quantity) / Units No."""
    input_rate = pd.to_numeric(input_rate, errors='coerce').fillna(0.0)
    quantity = pd.to_numeric(quantity, errors='coerce').fillna(0.0)
    units_no = pd.to_numeric(units_no, errors='coerce').fillna(1.0)
    units_no = units_no.apply(lambda x: 1.0 if x == 0 else x)
    return (input_rate * quantity) / units_no

def prepare_df_for_editor(df_original, pdf_column_mapping_rules, final_excel_column_order):
    """
    Prepares a DataFrame for st.data_editor with all required columns,
    including extracted, calculated, and user-input columns.
    All column names will be flattened strings.
    """
    temp_df = df_original.copy()
    temp_df.columns = [str(col).strip() for col in temp_df.columns]

    processed_df = pd.DataFrame(columns=final_excel_column_order)

    for excel_col_name in final_excel_column_order:
        found_in_pdf = False
        for pdf_target_col, rules in pdf_column_mapping_rules.items():
            if excel_col_name == pdf_target_col:
                for keyword in rules["keywords"]:
                    matching_original_col = next((
                        orig_col for orig_col in temp_df.columns if keyword.lower() in orig_col.lower()
                    ), None)

                    if matching_original_col and matching_original_col in temp_df.columns:
                        processed_df[excel_col_name] = temp_df[matching_original_col].reset_index(drop=True)
                        found_in_pdf = True
                        break
            if found_in_pdf:
                break

        if not found_in_pdf:
            series_length = len(temp_df) if not temp_df.empty else 0
            processed_df[excel_col_name] = pd.Series(dtype='object', index=range(series_length))

    if processed_df.empty:
        processed_df = pd.DataFrame(columns=final_excel_column_order, index=[0])

    for col in ["Quantity", "Govt Rate - Input", "Quoted Rate - Input"]:
        if col in processed_df.columns:
            processed_df[col] = pd.to_numeric(processed_df[col], errors='coerce').fillna(0.0)

    for col in ["Units No.", "Govt Rate - Total", "Quoted Rate - Total"]:
        if col in processed_df.columns:
            processed_df[col] = 0.0

    if "Units" not in processed_df.columns:
        processed_df["Units"] = ""
    if "Sr No." not in processed_df.columns:
        processed_df["Sr No."] = ""
    if "Items Description" not in processed_df.columns:
        processed_df["Items Description"] = ""

    return processed_df

def recalculate_editor_df_values(df):
    """
    Recalculates 'Units No.' and all 'Total Rate' columns
    based on user edits in 'Input Rate', 'Quantity', and 'Unit' columns.
    """
    df_copy = df.copy()

    if "Units" in df_copy.columns:
        new_units_no_from_units = df_copy["Units"].apply(convert_unit_to_number)
        df_copy["Units No."] = new_units_no_from_units

    df_copy["Units No."] = pd.to_numeric(df_copy["Units No."], errors='coerce').fillna(1.0)
    df_copy["Units No."] = df_copy["Units No."].apply(lambda x: 1.0 if x == 0 else x)

    for col in ["Quantity", "Govt Rate - Input", "Quoted Rate - Input"]:
        if col in df_copy.columns:
            df_copy[col] = pd.to_numeric(df_copy[col], errors='coerce').fillna(0.0)

    if "Govt Rate - Input" in df_copy.columns and "Quantity" in df_copy.columns and "Units No." in df_copy.columns:
        df_copy["Govt Rate - Total"] = calculate_total_rate(
            df_copy["Govt Rate - Input"], df_copy["Quantity"], df_copy["Units No."])

    if "Quoted Rate - Input" in df_copy.columns and "Quantity" in df_copy.columns and "Units No." in df_copy.columns:
        df_copy["Quoted Rate - Total"] = calculate_total_rate(
            df_copy["Quoted Rate - Input"], df_copy["Quantity"], df_copy["Units No."])

    return df_copy


# --- Streamlit UI ---
st.set_page_config(page_title="PDF Table Extractor and Editor", layout="wide")

st.title("📄 PDF Table Extractor & Editor")
st.markdown("""
Upload a PDF file to extract, edit, and consolidate tables. Calculated columns will update automatically.

*Important:* This app will now use `pdfplumber` for extraction. You may need to fine-tune extraction settings for best results on different PDFs.
""")

# --- SESSION STATE INITIALIZATION ---
final_display_excel_column_order = [
    "Sr No.", "Items Description", "Units", "Units No.", "Quantity",
    "Govt Rate - Input", "Govt Rate - Total", "Quoted Rate - Input", "Quoted Rate - Total"
]

if 'extraction_success' not in st.session_state:
    st.session_state.extraction_success = False

if 'single_combined_df' not in st.session_state:
    st.session_state.single_combined_df = pd.DataFrame(columns=final_display_excel_column_order)

# --- PDF Upload and Extraction Settings ---
uploaded_file = st.file_uploader("Upload your PDF file", type=["pdf"], key="pdf_uploader")

pages_input = st.text_input(
    "Enter pages to extract (e.g., '1,3-5'). Leave blank for all pages with tables.",
    value="",
    key="pages_input"
)

# --- Name Input for Excel Sheet (also used for summary) ---
name_input = st.text_input("Name (will appear in summary and Excel sheet)", value="Name:", key="name_input")

# --- Extraction Button ---
extract_button = st.button("Extract Tables for Editing", key="extract_button")

# --- Define PDF Column Mapping Rules ---
pdf_column_mapping_rules = {
    "Sr No.": {"keywords": ["sr.no", "sr no.", "sr no", "seriel number", "srno", "s.no", "serial no", "serial number",
                             "sr. no.", "sr.", "no."], "include_header_cell_in_data": False},
    "Items Description": {
        "keywords": ["items", "item description", "items description", "item name", "items name", "item", "description",
                     "item desc"], "include_header_cell_in_data": False},
    "Units": {"keywords": ["units", "unit"], "include_header_cell_in_data": False},
    "Quantity": {
        "keywords": ["quantity", "quantities", "estimated quantity", "estimated qty", "qty", "est. qty", "est qty",
                     "estimated quanity"], "include_header_cell_in_data": False},
    "Govt Rate - Input": {
        "keywords": ["est", "est rates", "est rate", "estimated rates", "estimated rate", "rate", "rates",
                     "market rate", "mkt rate", "mrkt rate", "market rates"], "include_header_cell_in_data": False},
    "Quoted Rate - Input": {"keywords": ["quoted rate", "quote rate", "quoted rates"],
                            "include_header_cell_in_data": False},
}

# --- Main Extraction and Display Logic ---
if extract_button:
    if not uploaded_file:
        st.warning("Please upload a PDF file first!")
    else:
        pdf_stream = io.BytesIO(uploaded_file.read())

        try:
            pages_to_extract = None
            if pages_input.strip() != "":
                # Convert '1,3-5' to [1, 3, 4, 5]
                pages_list = []
                for p_range in pages_input.split(','):
                    if '-' in p_range:
                        start, end = map(int, p_range.split('-'))
                        pages_list.extend(range(start, end + 1))
                    else:
                        pages_list.append(int(p_range))
                pages_to_extract = sorted(list(set(pages_list))) # Unique and sorted

            all_extracted_dfs_from_pdfplumber = []

            with st.spinner("Extracting tables using `pdfplumber`..."):
                with pdfplumber.open(pdf_stream) as pdf:
                    if pages_to_extract:
                        # Extract only specified pages
                        pages_iter = [pdf.pages[p - 1] for p in pages_to_extract if p <= len(pdf.pages)]
                    else:
                        # Iterate through all pages if no specific pages are given
                        pages_iter = pdf.pages

                    for page in pages_iter:
                        # pdfplumber's extract_tables method
                        # You might need to adjust table_settings for your specific PDFs
                        # 'snap_edges': False generally makes it more robust, 'line_scale' is like Camelot's
                        # 'vertical_strategy': 'lines' and 'horizontal_strategy': 'lines' often work well
                        tables_on_page = page.extract_tables(
                            table_settings={
                                "vertical_strategy": "lines",  # or "text" or "lines_strict" or "explicit_lines"
                                "horizontal_strategy": "lines", # or "text" or "lines_strict" or "explicit_lines"
                                "snap_tolerance": 3, # How close lines need to be to be considered snapped
                                "snap_edges": False,
                                "join_tolerance": 3,
                                "edge_min_length": 5, # Minimum length of a line to be considered
                                "min_words_horizontal": 1,
                                "min_words_vertical": 1,
                            }
                        )
                        for table_data in tables_on_page:
                            # Convert list of lists (table_data) to DataFrame
                            if table_data:
                                # First row is assumed to be header, rest is data
                                header = table_data[0]
                                data = table_data[1:]
                                df_temp = pd.DataFrame(data, columns=header)
                                all_extracted_dfs_from_pdfplumber.append(df_temp)

            if len(all_extracted_dfs_from_pdfplumber) == 0:
                st.warning("No tables were found on the specified pages using `pdfplumber`.")
                st.info("You may need to adjust the `table_settings` in the code for better extraction, or the PDF might not contain table structures `pdfplumber` can detect.")
                st.session_state.extraction_success = False
                st.stop()
            else:
                st.success(f"Successfully extracted {len(all_extracted_dfs_from_pdfplumber)} table(s) using `pdfplumber`.")

                # For simplicity, let's process the first extracted table to find header,
                # then apply that mapping to all tables.
                first_extracted_df_raw = all_extracted_dfs_from_pdfplumber[0].copy()
                # Clean up initial header row for matching
                first_extracted_df_raw.columns = [str(col).strip() for col in first_extracted_df_raw.columns]


                best_header_row_idx_first_page = -1 # This logic might need refinement for pdfplumber's raw output
                max_matches_first_page = 0
                header_column_map_from_first_page = {}

                if not all_extracted_dfs_from_pdfplumber: # Should have been caught by early warning
                    st.info("No tables to process after extraction.")
                    st.session_state.extraction_success = False
                    st.stop()

                # Get the first DataFrame and attempt to find headers
                df_first_extracted_table = all_extracted_dfs_from_pdfplumber[0].copy()

                # Clean up extracted header row values for matching
                # pdfplumber's extract_tables returns a list of lists.
                # The first sublist is usually the header row.
                # Let's make sure our header matching logic works with this.

                # Use the actual column names from the first extracted df for matching
                potential_header_row_values = [str(col).strip().lower() for col in df_first_extracted_table.columns.tolist()]

                current_matches = 0
                temp_header_col_map = {}

                for target_col_name, rules in pdf_column_mapping_rules.items():
                    for i, header_cell_content in enumerate(potential_header_row_values):
                        if any(keyword in header_cell_content for keyword in rules["keywords"]):
                            # Map the target column name (e.g., "Sr No.") to its index in the extracted DataFrame
                            temp_header_col_map[target_col_name] = df_first_extracted_table.columns[i] # Store the actual column name from pdfplumber
                            current_matches += 1
                            break # Move to next target_col_name

                if current_matches > 0: # If at least one header found
                    st.success(f"Headers identified in the first extracted table. Applying this structure.")
                    st.session_state.header_detection_failed = False
                    header_column_map_from_first_page = temp_header_col_map.copy()
                else:
                    st.info("Could not identify clear headers in the first extracted table. Displaying raw extraction and attempting best-effort mapping.")
                    st.session_state.header_detection_failed = True


                all_processed_dfs_for_concat = []

                for table_idx, df_raw_pdfplumber in enumerate(all_extracted_dfs_from_pdfplumber):
                    # For pdfplumber, the first row of each extracted table is its header.
                    # We'll use the header mapping derived from the FIRST table.
                    # If header detection failed, we just pass the raw data.

                    current_processed_df_data = {}

                    if st.session_state.header_detection_failed:
                        # Fallback: if header detection failed, just try to map existing columns if they match
                        editor_df = pd.DataFrame(columns=final_display_excel_column_order)
                        raw_cols_cleaned = [str(c).strip().lower() for c in df_raw_pdfplumber.columns.tolist()]
                        for final_col_name in final_display_excel_column_order:
                            found_col = False
                            for map_key, rules in pdf_column_mapping_rules.items():
                                if final_col_name == map_key:
                                    for keyword in rules['keywords']:
                                        if keyword in raw_cols_cleaned:
                                            try:
                                                col_index = raw_cols_cleaned.index(keyword)
                                                # Use the actual column name from df_raw_pdfplumber
                                                editor_df[final_col_name] = df_raw_pdfplumber[df_raw_pdfplumber.columns[col_index]].reset_index(drop=True)
                                                found_col = True
                                                break
                                            except ValueError:
                                                pass # Keyword not in current raw_cols_cleaned
                                if found_col: break
                            if not found_col:
                                editor_df[final_col_name] = pd.Series(dtype='object', index=range(len(df_raw_pdfplumber)))
                        # If df_raw_pdfplumber was empty or just header, ensure editor_df is not empty
                        if editor_df.empty and len(df_raw_pdfplumber) <=1:
                            editor_df = pd.DataFrame(columns=final_display_excel_column_order, index=[0])

                    else: # Headers were successfully identified from the first table
                        for target_col_name in final_display_excel_column_order:
                            original_col_name_from_pdfplumber = header_column_map_from_first_page.get(target_col_name)

                            if original_col_name_from_pdfplumber is not None and original_col_name_from_pdfplumber in df_raw_pdfplumber.columns:
                                col_data = df_raw_pdfplumber[original_col_name_from_pdfplumber].reset_index(drop=True)
                                current_processed_df_data[target_col_name] = col_data
                            else:
                                current_processed_df_data[target_col_name] = pd.Series(dtype='object', index=range(len(df_raw_pdfplumber)))

                        extracted_flat_df = pd.DataFrame(current_processed_df_data)

                        editor_df = prepare_df_for_editor(
                            extracted_flat_df,
                            pdf_column_mapping_rules, # Still useful for identifying column types based on content
                            final_display_excel_column_order
                        )

                    editor_df = recalculate_editor_df_values(editor_df)
                    all_processed_dfs_for_concat.append(editor_df)


                if all_processed_dfs_for_concat:
                    st.session_state.single_combined_df = pd.concat(all_processed_dfs_for_concat, ignore_index=True)
                    st.session_state.extraction_success = True
                else:
                    st.info("No tables could be processed with `pdfplumber` for display.")
                    st.session_state.extraction_success = False

        except Exception as e:
            st.error(f"An error occurred during PDF extraction: {e}")
            st.info("Please check your PDF file or page number input. `pdfplumber` relies on clear table structures. Error details: " + str(e))
            st.session_state.extraction_success = False

# --- Display Data Editor if Extraction was Successful and data exists ---
# (The rest of your code from here downwards remains exactly the same)
if 'extraction_success' in st.session_state and st.session_state.extraction_success and 'single_combined_df' in st.session_state and not st.session_state.single_combined_df.empty:
    st.subheader("Edit Extracted Tables")
    st.caption("Double-click a cell to edit. Press 'Enter' or click outside to see calculations update.")
    st.caption("Grayed out columns are calculated and not editable directly.")

    df_for_editor = st.session_state.single_combined_df.copy()

    column_config_dict = {
        "Units No.": st.column_config.NumberColumn("Units No.",
                                                   help="Calculated from Units, but can be manually overridden.",
                                                   format="%.2f", disabled=False),
        "Govt Rate - Total": st.column_config.NumberColumn("Govt Rate - Total",
                                                           help="Calculated from Govt Rate - Input * Quantity / Units No.",
                                                           format="%.2f", disabled=True),
        "Quoted Rate - Total": st.column_config.NumberColumn("Quoted Rate - Total",
                                                              help="Calculated from Quoted Rate - Input * Quantity / Units No.",
                                                              format="%.2f", disabled=True),

        "Units": st.column_config.TextColumn("Units", help="Enter unit (e.g., 'P-no/P Hour'). Affects 'Units No.'.",
                                             width="small"),
        "Quantity": st.column_config.NumberColumn("Quantity", help="Number of items or amount.", format="%.2f"),
        "Govt Rate - Input": st.column_config.NumberColumn("Govt Rate - Input", help="Enter the government input rate.",
                                                           format="%.2f"),
        "Quoted Rate - Input": st.column_config.NumberColumn("Quoted Rate - Input",
                                                              help="Enter the quoted input rate (for comparison).",
                                                              format="%.2f"),

        "Sr No.": st.column_config.TextColumn("Sr No.", disabled=True),
        "Items Description": st.column_config.TextColumn("Items Description", disabled=True, width="large"),
    }

    active_column_config = {}

    for col_name in final_display_excel_column_order:
        if col_name in df_for_editor.columns:
            if col_name in column_config_dict:
                active_column_config[col_name] = column_config_dict[col_name]
            else:
                active_column_config[col_name] = st.column_config.TextColumn(label=col_name, disabled=True)

    edited_df_from_widget = st.data_editor(
        df_for_editor,
        key="single_combined_data_editor",
        hide_index=True,
        num_rows="dynamic",
        column_config=active_column_config
    )

    recalculated_combined_df = recalculate_editor_df_values(edited_df_from_widget)
    st.session_state.single_combined_df = recalculated_combined_df

    st.markdown("---")

    if not st.session_state.single_combined_df.empty:
        combined_final_df = st.session_state.single_combined_df

        st.subheader("Summary")
        st.write(f"**Name:** {name_input}")

        total_govt_rate_sum = combined_final_df[
            "Govt Rate - Total"].sum() if "Govt Rate - Total" in combined_final_df.columns else 0
        total_quoted_rate_sum = combined_final_df[
            "Quoted Rate - Total"].sum() if "Quoted Rate - Total" in combined_final_df.columns else 0

        st.metric(label="Grand Total (Govt)", value=f"{total_govt_rate_sum:,.2f}")
        st.metric(label="Grand Total (Quoted)", value=f"{total_quoted_rate_sum:,.2f}")

        if total_govt_rate_sum > 0:
            rate_below_govt = 1 - (total_quoted_rate_sum / total_govt_rate_sum)
            st.metric(label="Rate Below From Govt Rate", value=f"{rate_below_govt:.2%}")
        else:
            st.info("Cannot calculate 'Rate Below From Govt Rate': Govt Grand Total is zero.")

        st.markdown("---")
        st.success("You can now download the combined, edited data as Excel.")

        excel_output_buffer = io.BytesIO()
        wb = Workbook()
        wb.remove(wb.active)

        sheet_name = "Combined Tables"
        ws = wb.create_sheet(title=sheet_name)

        col_to_excel_letter = {col_name: get_column_letter(idx + 1)
                               for idx, col_name in enumerate(final_display_excel_column_order)}

        initial_data_row = 4

        for _ in range(initial_data_row - 1):
            ws.append([])

        ws.append(final_display_excel_column_order)

        header_row_obj = ws[initial_data_row]
        for cell in header_row_obj:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

        ws.freeze_panes = get_column_letter(1) + str(initial_data_row + 1)

        for r_idx, row_data_series in combined_final_df.iterrows():
            excel_row_num = r_idx + initial_data_row + 1
            row_values_for_excel = []
            for col_name in final_display_excel_column_order:
                if col_name in ["Govt Rate - Total", "Quoted Rate - Total"]:
                    input_rate_col_name = ""
                    if col_name == "Govt Rate - Total":
                        input_rate_col_name = "Govt Rate - Input"
                    elif col_name == "Quoted Rate - Total":
                        input_rate_col_name = "Quoted Rate - Input"

                    input_rate_col_letter = col_to_excel_letter.get(input_rate_col_name)
                    quantity_col_letter = col_to_excel_letter.get("Quantity")
                    units_no_col_letter = col_to_excel_letter.get("Units No.")

                    if all(c in final_display_excel_column_order for c in
                           [input_rate_col_name, "Quantity", "Units No."]) and \
                            input_rate_col_letter and quantity_col_letter and units_no_col_letter:
                        formula_str = (
                            f"=(IFERROR(VALUE({input_rate_col_letter}{excel_row_num}),0)*IFERROR(VALUE({quantity_col_letter}{excel_row_num}),0))/"
                            f"(IF(OR(ISBLANK({units_no_col_letter}{excel_row_num}),IFERROR(VALUE({units_no_col_letter}{excel_row_num}),1)=0),1,IFERROR(VALUE({units_no_col_letter}{excel_row_num}),1)))"
                        )
                        row_values_for_excel.append(formula_str)
                    else:
                        row_values_for_excel.append(None)
                else:
                    cell_value = row_data_series.get(col_name)
                    row_values_for_excel.append(None if pd.isna(cell_value) else str(cell_value).strip())
            ws.append(row_values_for_excel)

        # Set column widths and wrap text
        for col_idx, column_header in enumerate(final_display_excel_column_order):
            col_letter = get_column_letter(col_idx + 1)
            if column_header == "Items Description":
                ws.column_dimensions[col_letter].width = 70
                for r in range(initial_data_row + 1, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx + 1).alignment = Alignment(wrap_text=True, vertical='top')
            else:
                ws.column_dimensions[col_letter].width = 12
                for r in range(initial_data_row + 1, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx + 1).alignment = Alignment(wrap_text=True, vertical='top')

        # Set row heights for wrap text
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = None

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=initial_data_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        current_data_end_row = ws.max_row

        govt_total_col_name = "Govt Rate - Total"
        quoted_total_col_name = "Quoted Rate - Total"

        if govt_total_col_name in final_display_excel_column_order:
            govt_total_col_letter = col_to_excel_letter.get(govt_total_col_name)
            if govt_total_col_letter:
                start_row_for_sum = initial_data_row + 1
                end_row_for_sum = current_data_end_row

                grand_total_label_row_govt = [None] * len(final_display_excel_column_order)
                if "Sr No." in final_display_excel_column_order:
                    grand_total_label_row_govt[final_display_excel_column_order.index("Sr No.")] = "Grand Total (Govt)"
                ws.append(grand_total_label_row_govt)
                row_num_for_govt_grand_total = ws.max_row

                grand_total_formula_cell_govt = ws.cell(row=row_num_for_govt_grand_total,
                                                         column=final_display_excel_column_order.index(
                                                             govt_total_col_name) + 1)
                grand_total_formula_cell_govt.value = f"=SUM({govt_total_col_letter}{start_row_for_sum}:{govt_total_col_letter}{end_row_for_sum})"
                for cell in ws[row_num_for_govt_grand_total]:
                    cell.font = Font(bold=True)
                    cell.border = thin_border

        if quoted_total_col_name in final_display_excel_column_order:
            quoted_total_col_letter = col_to_excel_letter.get(quoted_total_col_name)
            if quoted_total_col_letter:
                start_row_for_sum = initial_data_row + 1
                end_row_for_sum = current_data_end_row

                grand_total_label_row_quoted = [None] * len(final_display_excel_column_order)
                if "Sr No." in final_display_excel_column_order:
                    grand_total_label_row_quoted[
                        final_display_excel_column_order.index("Sr No.")] = "Grand Total (Quoted)"
                ws.append(grand_total_label_row_quoted)
                row_num_for_quoted_grand_total = ws.max_row

                grand_total_formula_cell_quoted = ws.cell(row=row_num_for_quoted_grand_total,
                                                           column=final_display_excel_column_order.index(
                                                               quoted_total_col_name) + 1)
                grand_total_formula_cell_quoted.value = f"=SUM({quoted_total_col_letter}{start_row_for_sum}:{quoted_total_col_letter}{end_row_for_sum})"
                for cell in ws[row_num_for_quoted_grand_total]:
                    cell.font = Font(bold=True)
                    cell.border = thin_border

        if name_input:
            ws['A1'] = name_input
            ws['A1'].font = Font(bold=True)
            ws['A1'].border = thin_border

        govt_rate_total_col_letter = col_to_excel_letter.get(govt_total_col_name)
        quoted_rate_total_col_letter = col_to_excel_letter.get(quoted_total_col_name)

        if govt_rate_total_col_letter and quoted_rate_total_col_letter and 'row_num_for_govt_grand_total' in locals() and 'row_num_for_quoted_grand_total' in locals():
            ws[f'{govt_rate_total_col_letter}1'] = 'Rate Below From Govt Rate'
            ws[f'{govt_rate_total_col_letter}1'].font = Font(bold=True)
            ws[f'{govt_rate_total_col_letter}1'].alignment = Alignment(horizontal='right')
            ws[f'{govt_rate_total_col_letter}1'].border = thin_border
            ws[f'{govt_rate_total_col_letter}1'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                                                                     fill_type="solid")

            formula_cell_loc = f'{quoted_rate_total_col_letter}1'
            formula_summary_str = (
                f'=IFERROR(1-(IFERROR(VALUE({quoted_rate_total_col_letter}{row_num_for_quoted_grand_total}),0)/'
                f'IF(IFERROR(VALUE({govt_rate_total_col_letter}{row_num_for_govt_grand_total}),0)=0,1,IFERROR(VALUE({govt_rate_total_col_letter}{row_num_for_govt_grand_total}),0))),0)'
            )
            ws[formula_cell_loc] = formula_summary_str
            ws[formula_cell_loc].number_format = '0.00%'
            ws[formula_cell_loc].font = Font(bold=True)
            ws[formula_cell_loc].border = thin_border
            ws[formula_cell_loc].fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        else:
            st.warning("Could not find required columns for 'Rate Below From Govt Rate' summary.")

        wb.save(excel_output_buffer)
        excel_output_buffer.seek(0)

        st.download_button(
            label="Download Edited Tables as Excel (.xlsx)",
            data=excel_output_buffer.getvalue(),
            file_name="edited_tables_with_formulas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("No tables to display or download after processing.")

st.markdown("---")
st.markdown("Developed with ❤ using Streamlit, PDFPlumber, Pandas, OpenPyXL.") # Updated credit